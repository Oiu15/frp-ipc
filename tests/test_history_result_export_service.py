import csv
import json
import math
import shutil
import unittest
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from services.history_result_export_service import (
    DETECTION_SUMMARY_COLUMNS,
    HistoryResultExportService,
)


SUMMARY_HEADER = [
    "date",
    "start_time",
    "end_time",
    "serial",
    "run_id",
    "recipe_name",
    "od_end_off_mm",
    "id_end_off_mm",
    "od_range_mm",
    "od_mean_mm",
    "id_mean_mm",
    "status",
]

SECTION_HEADER = [
    "serial",
    "run_id",
    "start_time",
    "end_time",
    "duration_s",
    "section_idx",
    "z_pos_mm",
    "od_avg_mm",
    "id_avg_mm",
    "od_pp_rob_mm",
    "id_round_fit_mm",
    "concentricity_mm",
]


def _sheet_row_values(ws: Any, row_index: int) -> list[Any]:
    return [cell.value for cell in ws[row_index]]


class HistoryResultExportServiceTest(unittest.TestCase):
    def _case_root(self, name: str) -> Path:
        root = Path(".test-artifacts") / name
        shutil.rmtree(root, ignore_errors=True)
        root.mkdir(parents=True, exist_ok=True)
        return root / "FRP_IPC"

    def _write_run(
        self,
        app_root: Path,
        *,
        date: str,
        serial: str,
        run_id: str | None = None,
        completed: bool | None = True,
        status: str = "DONE",
        sections: int = 5,
        write_meta: bool = True,
        write_section: bool = True,
        write_summary_row: bool = True,
        od_mean: str = "101.500",
        id_mean: str = "51.500",
    ) -> Path:
        run_id = run_id or f"run-{serial}"
        day_dir = app_root / "exports" / date
        run_dir = day_dir / serial
        run_dir.mkdir(parents=True, exist_ok=True)

        if write_meta:
            meta = {
                "serial": serial,
                "run_id": run_id,
                "start_time": f"{date} 08:00:00",
                "status": status,
                "completed_sections": sections,
                "expected_sections": 5,
                "recipe": {"name": "recipe-a"},
            }
            if completed is not None:
                meta["completed"] = completed
            (run_dir / "meta.json").write_text(json.dumps(meta, ensure_ascii=False), encoding="utf-8")

        if write_section:
            with open(run_dir / "section_results.csv", "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(SECTION_HEADER)
                for idx in range(1, sections + 1):
                    writer.writerow([
                        serial,
                        run_id,
                        f"{date} 08:00:00",
                        f"{date} 08:01:00",
                        "60.000",
                        idx,
                        idx * 100,
                        100 + idx,
                        50 + idx,
                        0.10 * idx,
                        0.20 * idx,
                        0.30 * idx,
                    ])

        if write_summary_row:
            summary_path = day_dir / "summary.csv"
            new_file = not summary_path.exists()
            with open(summary_path, "a", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                if new_file:
                    writer.writerow(SUMMARY_HEADER)
                writer.writerow([
                    date,
                    "08:00:00",
                    "08:01:00",
                    serial,
                    run_id,
                    "recipe-a",
                    "1.234",
                    "2.345",
                    "3.456",
                    od_mean,
                    id_mean,
                    status,
                ])

        return run_dir

    def test_lists_only_exportable_complete_history_entries(self) -> None:
        app_root = self._case_root("history_export_list")
        self._write_run(app_root, date="2025-01-01", serial="good-001")
        self._write_run(app_root, date="2025-01-01", serial="legacy-002", completed=None, status="DONE")
        self._write_run(app_root, date="2025-01-01", serial="missing-meta", write_meta=False)
        self._write_run(app_root, date="2025-01-01", serial="missing-section", write_section=False)
        self._write_run(app_root, date="2025-01-01", serial="not-complete", completed=False, status="STOP")
        self._write_run(app_root, date="2025-01-01", serial="short-sections", sections=4)
        self._write_run(app_root, date="2025-01-01", serial="no-summary-row", write_summary_row=False)

        entries = HistoryResultExportService(app_root_dir=app_root).list_exportable_entries()

        self.assertEqual([entry.serial for entry in entries], ["good-001", "legacy-002"])

    def test_exports_selected_entries_with_required_column_order_and_sequence(self) -> None:
        app_root = self._case_root("history_export_xlsx")
        self._write_run(app_root, date="2025-01-01", serial="serial-a", od_mean="111.111", id_mean="55.555")
        self._write_run(app_root, date="2025-01-02", serial="serial-b", od_mean="222.222", id_mean="66.666")
        service = HistoryResultExportService(app_root_dir=app_root)
        entries_by_serial = {entry.serial: entry for entry in service.list_exportable_entries()}
        output_path = app_root / "manual" / "检测数据汇总.xlsx"

        service.export_detection_summary(
            [entries_by_serial["serial-b"], entries_by_serial["serial-a"]],
            output_path,
        )

        wb = load_workbook(output_path, data_only=True)
        ws = wb["检测数据汇总"]
        header = _sheet_row_values(ws, 1)
        row1 = _sheet_row_values(ws, 2)
        row2 = _sheet_row_values(ws, 3)

        self.assertEqual(header, list(DETECTION_SUMMARY_COLUMNS))
        self.assertEqual(row1[0], 1)
        self.assertEqual(row1[1], "serial-b")
        self.assertEqual(row2[0], 2)
        self.assertEqual(row2[1], "serial-a")
        self.assertTrue(math.isclose(float(row1[2]), 222.222))
        self.assertTrue(math.isclose(float(row1[8]), 1.234))
        self.assertTrue(math.isclose(float(row1[9]), 3.456))
        self.assertTrue(math.isclose(float(row1[10]), 66.666))
        self.assertTrue(math.isclose(float(row1[16]), 2.345))
        self.assertTrue(math.isclose(float(row1[17]), 0.3))

    def test_export_falls_back_to_section_average_when_summary_mean_missing(self) -> None:
        app_root = self._case_root("history_export_mean_fallback")
        self._write_run(app_root, date="2025-01-01", serial="fallback", od_mean="", id_mean="")
        service = HistoryResultExportService(app_root_dir=app_root)
        output_path = app_root / "检测数据汇总.xlsx"

        service.export_detection_summary(service.list_exportable_entries(), output_path)

        wb = load_workbook(output_path, data_only=True)
        ws = wb["检测数据汇总"]
        row = _sheet_row_values(ws, 2)

        self.assertTrue(math.isclose(float(row[2]), 103.0))
        self.assertTrue(math.isclose(float(row[10]), 53.0))


if __name__ == "__main__":
    unittest.main()
